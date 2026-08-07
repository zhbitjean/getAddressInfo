from __future__ import annotations

import io
import zipfile
import pytest
import app as app_module
from datetime import datetime
from requests.structures import CaseInsensitiveDict

from openpyxl import Workbook, load_workbook

from building_classes import BUILDING_CLASS_DESCRIPTIONS, format_building_class

from nyc_data import (
    BUILDING_FOOTPRINTS_URL,
    HPD_BUILDINGS_URL,
    HPD_VIOLATIONS_URL,
    GEOSEARCH_URL,
    LEGACY_CO_CONTENT_URL,
    LEGACY_CO_POST_URL,
    LEGACY_CO_URL,
    PLUTO_URL,
    NYCPropertyClient,
    PropertyRecord,
)
from workbook_io import DETAIL_COLUMNS, REPORT_ROWS, build_result_zip, read_addresses


def make_input() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Input"
    sheet.append(["Jobsite", "Borough"])
    sheet.append(["35 EUCLID AVENUE, Brooklyn, NY, 11208", "Brooklyn"])
    sheet.append(["220 LINCOLN ROAD, Brooklyn, NY, 11225", "Brooklyn"])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_read_addresses():
    addresses, source = read_addresses(make_input())
    assert len(addresses) == 2
    assert addresses[0].startswith("35 EUCLID")
    assert source == "Input!A1"


def test_read_addresses_uses_current_input_and_ignores_dated_tabs():
    workbook = Workbook()
    dated = workbook.active
    dated.title = "20260805"
    dated.append(["Jobsite"])
    dated.append(["OLD ADDRESS"])
    current = workbook.create_sheet("current_input")
    current.append(["Jobsite", "Measurement Date"])
    current.append(["643 Shepherd Ave, Brooklyn, NY 11208", "08/06/2026"])
    current.append(["191 19th St, Brooklyn, NY 11232", "08/06/2026"])
    workbook.create_sheet("20260804").append(["Jobsite"])
    output = io.BytesIO()
    workbook.save(output)

    addresses, source = read_addresses(output.getvalue())

    assert addresses == [
        "643 Shepherd Ave, Brooklyn, NY 11208",
        "191 19th St, Brooklyn, NY 11232",
    ]
    assert source == "current_input!A1"


def test_multitab_workbook_requires_current_input_tab():
    workbook = Workbook()
    workbook.active.title = "20260805"
    workbook.active.append(["Jobsite"])
    workbook.active.append(["OLD ADDRESS"])
    workbook.create_sheet("20260804").append(["Jobsite"])
    output = io.BytesIO()
    workbook.save(output)

    with pytest.raises(ValueError, match="current_input"):
        read_addresses(output.getvalue())

def test_default_http_session_retries_transient_get_failures():
    client = NYCPropertyClient()
    retry = client.session.get_adapter("https://").max_retries

    assert retry.total == 3
    assert retry.read == 3
    assert retry.connect == 3
    assert 429 in retry.status_forcelist
    assert 503 in retry.status_forcelist
    assert retry.allowed_methods == frozenset({"GET"})


def test_output_zip_name_contains_timestamp():
    name = app_module._timestamped_zip_name(
        "My Input",
        "results",
        datetime(2026, 8, 7, 14, 5, 9),
    )

    assert name == "My_Input_results_20260807_140509_000.zip"

def test_result_zip_contains_workbook_and_documents():
    record = PropertyRecord(
        input_address="35 EUCLID AVENUE, Brooklyn, NY, 11208",
        matched_address="35 EUCLID AVENUE",
        bbl="3041680016",
        bin="3094201",
        land_use="One & Two Family Buildings",
        lot_area="2760",
        co="Yes",
        co_download_status="Downloaded",
    )
    payload = build_result_zip([record], {record.input_address: [("CO_test.pdf", b"%PDF-test")]}, "Input!A1")
    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        assert "property_results.xlsx" in archive.namelist()
        assert any(name.endswith("CO_test.pdf") for name in archive.namelist())
        assert any(name.endswith("property_information.txt") for name in archive.namelist())
        workbook = load_workbook(io.BytesIO(archive.read("property_results.xlsx")), data_only=True)
        assert workbook["Results"]["A2"].value.startswith("35 EUCLID")
        assert workbook["Report View"]["B1"].value.startswith("35 EUCLID")


def test_lookup_preserves_geosearch_bin_when_pluto_omits_it():
    client = NYCPropertyClient()
    client._geocode = lambda _address: {
        "properties": {
            "label": "35 EUCLID AVENUE",
            "addendum": {"pad": {"bbl": "3041050020", "bin": "3090923"}},
        }
    }
    client._pluto = lambda *_args: {"address": "35 EUCLID AVENUE"}
    client._add_hpd = lambda *_args: None
    captured = {}
    client._add_co = lambda record, _files: captured.update(bin=record.bin)

    record, _files = client.lookup("35 EUCLID AVENUE, Brooklyn, NY")

    assert record.bin == "3090923"
    assert captured["bin"] == "3090923"

def test_result_zip_lists_failed_co_and_retry_links():
    record = PropertyRecord(
        input_address="48 WEST 22 STREET, MANHATTAN",
        bin="1015548",
        co="Yes",
        co_download_status="Failed: HTTP 403",
        co_retry_url="https://a810-bisweb.nyc.gov/bisweb/COsByLocationServlet?requestid=2&allbin=1015548",
    )

    payload = build_result_zip([record], {}, "retry-test")

    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        assert "failed_co_downloads.csv" in archive.namelist()
        assert "retry_failed_co_downloads.html" not in archive.namelist()
        csv_text = archive.read("failed_co_downloads.csv").decode("utf-8-sig")
        assert "HTTP 403" in csv_text
        assert "1015548" in csv_text
        assert "/retry-co?" in csv_text

def test_main_page_displays_failed_co_and_serves_result_zip():
    original_client = app_module.NYCPropertyClient

    class FakeClient:
        def lookup(self, address):
            record = PropertyRecord(
                input_address=address,
                matched_address=address,
                bin="1015548",
                co="Yes",
                co_download_status="Failed: BIS access denied",
                co_retry_url=(
                    "https://a810-bisweb.nyc.gov/bisweb/"
                    "COsByLocationServlet?requestid=2&allbin=1015548"
                ),
            )
            return record, []

    app_module.NYCPropertyClient = FakeClient
    try:
        client = app_module.app.test_client()
        response = client.post(
            "/process",
            data={
                "workbook": (io.BytesIO(make_input()), "input.xlsx"),
                "column": "Jobsite",
            },
            content_type="multipart/form-data",
        )
    finally:
        app_module.NYCPropertyClient = original_client

    assert response.status_code == 200
    page = response.get_data(as_text=True)
    assert "CO 下载失败（2）" in page
    assert "BIS access denied" in page
    assert "重试自动下载" in page
    assert "打开 BIS 手动下载" in page
    assert "retry_failed_co_downloads.html" not in page
    assert "Last query finished at:" in page
    assert "input_results_" in page

    import re
    match = re.search(r'href="(/download-result/[^"]+)"', page)
    assert match
    download = client.get(match.group(1))
    assert download.status_code == 200
    with zipfile.ZipFile(io.BytesIO(download.data)) as archive:
        assert "failed_co_downloads.csv" in archive.namelist()
        assert "retry_failed_co_downloads.html" not in archive.namelist()

    home_page = client.get("/").get_data(as_text=True)
    assert "Last completed query" in home_page
    assert "Download this ZIP again" in home_page
    assert "input_results_" in home_page

def test_co_download_fields_are_the_last_two_workbook_rows():
    assert DETAIL_COLUMNS[-2:] == [
        ("CO Download Status", "co_download_status"),
        ("CO Retry URL", "co_retry_url"),
    ]
    assert REPORT_ROWS[-2:] == [
        ("CO download status", "co_download_status"),
        ("CO retry URL", "co_retry_url"),
    ]


def test_boolean_document_rows_are_removed_and_missing_dates_show_no():
    detail_keys = [key for _label, key in DETAIL_COLUMNS]
    report_keys = [key for _label, key in REPORT_ROWS]
    record = PropertyRecord(input_address="TEST")

    assert "co" not in detail_keys
    assert "historical_image_cards" not in detail_keys
    assert "co" not in report_keys
    assert "historical_image_cards" not in report_keys
    assert record.co_date == "No"
    assert record.historical_image_cards_date == "No"

def test_official_building_class_mapping_is_complete_and_readable():
    assert len(BUILDING_CLASS_DESCRIPTIONS) >= 220
    assert format_building_class("B1") == "Two Family Dwellings - Brick (B1)"
    assert format_building_class("A5") == "One Family Dwellings - Attached or Semi-Detached (A5)"
    assert format_building_class("S2") == (
        "Residence (Multiple Use) - Primarily Two Family with One Store or Office (S2)"
    )
    assert format_building_class("B3") == "Converted From One Family (B3)"
    assert format_building_class("A4") == "One Family Dwellings - City Residence (A4)"
    assert format_building_class("C1") == "Walk-Up Apartments - Over Six Families without Stores (C1)"
    assert format_building_class("not-a-code") == "NOT-A-CODE"


def test_historic_district_is_reported_as_landmark():
    record = PropertyRecord(input_address="1073 Prospect Pl, Brooklyn, NY 11213")

    NYCPropertyClient._apply_pluto(
        record,
        {
            "address": "1073 PROSPECT PLACE",
            "histdist": "Crown Heights North III Historic District",
        },
    )

    assert record.landmark_status == (
        "L - LANDMARK (Crown Heights North III Historic District)"
    )


def test_individual_landmark_takes_precedence_over_historic_district():
    record = PropertyRecord(input_address="TEST")

    NYCPropertyClient._apply_pluto(
        record,
        {
            "landmark": "Example Individual Landmark",
            "histdist": "Example Historic District",
        },
    )

    assert record.landmark_status == "L - LANDMARK (Example Individual Landmark)"

def test_pluto_building_class_code_is_expanded():
    record = PropertyRecord(input_address="TEST")
    NYCPropertyClient._apply_pluto(record, {"bldgclass": "B1"})

    assert record.building_class == "Two Family Dwellings - Brick (B1)"

def test_address_normalization_for_exact_pluto_lookup():
    normalize = NYCPropertyClient._address_and_zip

    assert normalize("571 Leonard St, Brooklyn, NY 11222") == ("571 LEONARD STREET", "11222")
    assert normalize("191 19th St, Brooklyn, NY 11232") == ("191 19 STREET", "11232")


def test_geocode_prefers_exact_pluto_address_and_normalizes_bbl():
    client = NYCPropertyClient()
    calls = []

    def fake_json(url, params):
        calls.append((url, params))
        assert url == PLUTO_URL
        return [{
            "address": "22 EAST 13 STREET",
            "bbl": "1005700055.00000000",
            "bin": "1000001",
        }]

    client._json = fake_json

    feature = client._geocode("22 East 13 St, New York, NY 10003")

    assert feature["properties"]["label"] == "22 EAST 13 STREET"
    assert feature["properties"]["bbl"] == "1005700055"
    assert calls == [(
        PLUTO_URL,
        {"$where": "upper(address)='22 EAST 13 STREET' AND zipcode='10003'", "$limit": 10},
    )]


def test_geocode_uses_geosearch_when_pluto_has_no_exact_match():
    client = NYCPropertyClient()
    calls = []

    def fake_json(url, params):
        calls.append(url)
        if url == PLUTO_URL:
            return []
        assert url == GEOSEARCH_URL
        return {"features": [{"properties": {"label": "FUZZY MATCH"}}]}

    client._json = fake_json

    feature = client._geocode("Unusual address without ZIP")

    assert feature["properties"]["label"] == "FUZZY MATCH"
    assert calls == [PLUTO_URL, GEOSEARCH_URL]


def test_building_footprints_supplies_bin_when_pluto_omits_it():
    client = NYCPropertyClient()
    record = PropertyRecord(
        input_address="191 19th St, Brooklyn, NY 11232",
        bbl="3006340069",
    )
    captured = {}

    def fake_json(url, params):
        captured.update(url=url, params=params)
        return [{"bin": "3009112", "base_bbl": "3006340069"}]

    client._json = fake_json
    client._add_bin_from_footprints(record)

    assert record.bin == "3009112"
    assert captured == {
        "url": BUILDING_FOOTPRINTS_URL,
        "params": {"$where": "base_bbl='3006340069'", "$limit": 10},
    }

def test_hpd_supplies_bin_when_pluto_does_not():
    client = NYCPropertyClient()
    record = PropertyRecord(input_address="571 Leonard St, Brooklyn, NY 11222")

    client._json = lambda url, _params: (
        [{"buildingid": "", "bin": "3065517"}]
        if url == HPD_BUILDINGS_URL
        else [{"total": "0"}]
    )

    client._add_hpd(record, "3", "2647", "27", [])

    assert record.bin == "3065517"
    assert "bin=3065517" in record.dob_url

def test_pluto_query_uses_numeric_bbl_fields():
    client = NYCPropertyClient()
    captured = {}

    def fake_json(url, params):
        captured.update(url=url, params=params)
        return [{"address": "TEST ADDRESS"}]

    client._json = fake_json

    assert client._pluto("3", "1055", "4") == {"address": "TEST ADDRESS"}
    assert PLUTO_URL.endswith("/64uk-42ks.json")
    assert captured == {
        "url": PLUTO_URL,
        "params": {"$where": "borocode=3 AND block=1055 AND lot=4", "$limit": 5},
    }


def test_hpd_file_selection_prefers_newest_date():
    key = NYCPropertyClient._file_selection_key
    candidates = [
        ("I-Card_900.pdf", "2026-06-01"),
        ("I-Card_100.pdf", "2026-07-01"),
    ]

    assert max(candidates, key=lambda item: key(*item))[0] == "I-Card_100.pdf"


def test_hpd_file_selection_uses_largest_numeric_filename_for_date_ties():
    key = NYCPropertyClient._file_selection_key
    candidates = [
        ("I-Card_9.pdf", "2026-07-01"),
        ("I-Card_10.pdf", "2026-07-01"),
        ("I-Card_2.pdf", "2026-07-01"),
    ]

    assert max(candidates, key=lambda item: key(*item))[0] == "I-Card_10.pdf"


def test_hpd_download_uses_current_public_api_with_building_id():
    client = NYCPropertyClient()
    record = PropertyRecord(
        input_address="155 Stuyvesant Ave, Brooklyn, NY 11221",
        matched_address="155 Stuyvesant Ave Brooklyn NY 11221",
    )
    client._json = lambda url, _params: (
        [{"buildingid": "378808", "registrationid": "734744"}]
        if url == HPD_BUILDINGS_URL
        else [{"total": "0"}]
    )
    requested = []
    client._download_hpd_image_card = lambda building_id: (
        requested.append(building_id)
        or ("Icard_734744.pdf", b"%PDF-current-api", "03/30/2008")
    )
    files = []

    client._add_hpd(record, "3", "1777", "54", files)

    assert requested == ["378808"]
    assert files == [("155_Stuyvesant_Ave_Brooklyn_NY_11221_i-card.pdf", b"%PDF-current-api")]
    assert record.historical_image_cards == "Yes"
    assert record.historical_image_cards_date == "03/30/2008"


def test_hpd_document_api_decodes_pdf():
    class JsonResponse:
        def __init__(self, payload):
            self.payload = payload

        def raise_for_status(self):
            pass

        def json(self):
            return self.payload

    class ApiSession:
        headers = CaseInsensitiveDict()

        def post(self, url, **kwargs):
            assert url.endswith("/Apim/token")
            return JsonResponse({"token": "public-token"})

        def get(self, url, **kwargs):
            assert kwargs["headers"] == {"ApiKey": "Bearer public-token"}
            if "historicimage/list" in url:
                return JsonResponse({
                    "responseData": [{
                        "imageSeqNo": 734744,
                        "docId": 4241075,
                        "docTypeId": 53,
                        "subDocTypeId": 73,
                        "dateTaken": "03/30/2008 00:00:00",
                        "fileType": "pdf",
                    }]
                })
            assert url.endswith("/734744/4241075/53/73")
            import base64
            return JsonResponse({
                "responseData": {
                    "documentBytes": base64.b64encode(b"%PDF-api").decode("ascii")
                }
            })

    result = NYCPropertyClient(session=ApiSession())._download_hpd_image_card("378808")

    assert result == ("Icard_734744.pdf", b"%PDF-api", "03/30/2008")

def test_hpd_download_uses_address_lookup_registration_id_without_building_id():
    client = NYCPropertyClient()
    record = PropertyRecord(
        input_address="155 Stuyvesant Ave, Brooklyn, NY 11221",
        matched_address="155 Stuyvesant Ave Brooklyn NY 11221",
    )
    files = []
    requested_urls = []

    def fake_json(url, _params):
        if url == HPD_BUILDINGS_URL:
            return [{
                "registrationid": "734744",
                "laststatusdate": "2026-07-01",
            }]
        assert url == HPD_VIOLATIONS_URL
        return [{"total": "0"}]

    def fake_download(url):
        requested_urls.append(url)
        return b"%PDF-test"

    client._json = fake_json
    client._download_pdf = fake_download
    client._add_hpd(record, "3", "1777", "54", files)

    assert requested_urls == ["https://hpdonline.nyc.gov/HPDonline/PDFs/Icard_734744.pdf"]
    assert files == [("155_Stuyvesant_Ave_Brooklyn_NY_11221_i-card.pdf", b"%PDF-test")]


def test_hpd_candidates_keep_newest_date_then_largest_registration_id():
    client = NYCPropertyClient()
    record = PropertyRecord(input_address="TEST ADDRESS", matched_address="TEST ADDRESS")
    files = []
    requested_urls = []

    def fake_json(url, _params):
        if url == HPD_BUILDINGS_URL:
            return [
                {"registrationid": "900", "laststatusdate": "2026-06-01"},
                {"registrationid": "100", "laststatusdate": "2026-07-01"},
                {"registrationid": "101", "laststatusdate": "2026-07-01"},
            ]
        return [{"total": "0"}]

    def fake_download(url):
        requested_urls.append(url)
        return b"%PDF-test"

    client._json = fake_json
    client._download_pdf = fake_download
    client._add_hpd(record, "3", "1", "1", files)

    assert requested_urls[0].endswith("/Icard_101.pdf")


class FakeResponse:
    def __init__(self, content=b"", status_code=200, url="https://example.test/"):
        self.content = content
        self.status_code = status_code
        self.url = url
        self.headers = {"content-type": "text/html"}

    def raise_for_status(self):
        if self.status_code >= 400:
            from requests import HTTPError

            raise HTTPError(f"HTTP {self.status_code}")


def test_co_download_accepts_direct_pdf_anchor():
    listing = b"""
        <a href="CofoJobDocumentServlet?cofomatadata5=M000085679.PDF">
          M000085679.PDF
        </a>
    """

    class FakeSession:
        headers = CaseInsensitiveDict()

        def get(self, url, **kwargs):
            assert url == LEGACY_CO_URL
            assert kwargs["params"]["requestid"] == 2
            return FakeResponse(listing, url=f"{LEGACY_CO_URL}?requestid=2&allbin=1015548")

    client = NYCPropertyClient(session=FakeSession())
    client._json = lambda *_args, **_kwargs: []
    requested_urls = []
    client._download_pdf = lambda url: requested_urls.append(url) or b"%PDF-anchor"
    record = PropertyRecord(input_address="48 West 22 Street", bin="1015548")
    files = []

    client._add_co(record, files)

    assert requested_urls == [
        "https://a810-bisweb.nyc.gov/bisweb/CofoJobDocumentServlet?cofomatadata5=M000085679.PDF"
    ]
    assert files == [("M000085679.PDF", b"%PDF-anchor")]
    assert record.co == "Yes"
    assert record.co_download_status == "Downloaded"
    assert record.co_retry_url.endswith("requestid=2&allbin=1015548")

def test_co_failure_is_visible_and_retryable():
    class BlockedSession:
        headers = CaseInsensitiveDict()

        def get(self, url, **_kwargs):
            assert url == LEGACY_CO_URL
            return FakeResponse(status_code=403)

    client = NYCPropertyClient(session=BlockedSession())
    client._json = lambda *_args, **_kwargs: []
    record = PropertyRecord(input_address="48 West 22 Street", bin="1015548")

    client._add_co(record, [])

    assert record.co_download_status.startswith("Failed:")
    assert "403" in record.co_download_status
    assert record.co_retry_url.endswith("requestid=2&allbin=1015548")

def test_co_access_denied_page_with_http_200_is_a_failure():
    class BlockedSession:
        headers = CaseInsensitiveDict()

        def get(self, _url, **_kwargs):
            return FakeResponse(b"<html><h1>Access Denied</h1></html>", status_code=200)

    client = NYCPropertyClient(session=BlockedSession())
    client._json = lambda *_args, **_kwargs: []
    record = PropertyRecord(input_address="48 West 22 Street", bin="1015548")

    client._add_co(record, [])

    assert record.co_download_status == "Failed: BIS access denied"

def test_co_download_falls_back_to_direct_content_servlet_when_wrapper_post_fails():
    listing = b"""
        <form action="CofoJobDocumentServlet" method="post">
          <input name="passcofonumber" value="B000191077.PDF">
          <input name="requestid" value="2">
          <input name="cofomatadata1" value="COFO">
          <input name="cofomatadata2" value="B">
          <input name="cofomatadata3" value="000">
          <input name="cofomatadata4" value="191000">
          <input name="cofomatadata5" value="B000191077.PDF">
        </form>
    """

    class FakeSession:
        headers = CaseInsensitiveDict()

        def get(self, url, **kwargs):
            assert url == LEGACY_CO_URL
            return FakeResponse(listing, url=f"{LEGACY_CO_URL}?requestid=1&allbin=3007398")

        def post(self, url, **kwargs):
            assert url == LEGACY_CO_POST_URL
            return FakeResponse(status_code=403)

    client = NYCPropertyClient(session=FakeSession())
    client._json = lambda *_args, **_kwargs: []
    requested = {}

    def fake_direct(url, params):
        requested.update(url=url, params=params)
        return b"%PDF-direct"

    client._download_pdf_with_params = fake_direct
    record = PropertyRecord(input_address="318 President St", bin="3007398")
    files = []
    client._add_co(record, files)

    assert files == [("B000191077.PDF", b"%PDF-direct")]
    assert requested["url"] == LEGACY_CO_CONTENT_URL
    assert requested["params"]["cofomatadata5"] == "B000191077.PDF"
    assert "requestid" not in requested["params"]
    assert record.co == "Yes"


def test_co_download_accepts_object_viewer_url():
    listing = b"""
        <form action="CofoJobDocumentServlet">
          <input name="passcofonumber" value="M000052537.PDF">
          <input name="cofomatadata1" value="COFO">
          <input name="cofomatadata5" value="M000052537.PDF">
        </form>
    """
    wrapper = b'<object data="CofoDocumentContentServlet?cofomatadata5=M000052537.PDF"></object>'

    class FakeSession:
        headers = CaseInsensitiveDict()

        def get(self, _url, **_kwargs):
            return FakeResponse(listing, url=LEGACY_CO_URL)

        def post(self, _url, **_kwargs):
            return FakeResponse(wrapper)

    client = NYCPropertyClient(session=FakeSession())
    client._json = lambda *_args, **_kwargs: []
    requested_urls = []
    client._download_pdf = lambda url: requested_urls.append(url) or b"%PDF-object"
    record = PropertyRecord(input_address="323 East 14 St", bin="1020397")
    files = []
    client._add_co(record, files)

    assert requested_urls == [
        "https://a810-bisweb.nyc.gov/bisweb/CofoDocumentContentServlet?cofomatadata5=M000052537.PDF"
    ]
    assert files == [("M000052537.PDF", b"%PDF-object")]
