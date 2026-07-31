from __future__ import annotations

import csv
import io
import itertools
import re
import zipfile
from pathlib import Path
from typing import Iterable
from urllib.parse import urlencode

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from nyc_data import PropertyRecord


DETAIL_COLUMNS = [
    ("Input Address", "input_address"),
    ("Matched Address", "matched_address"),
    ("Status", "status"),
    ("Error", "error"),
    ("BBL", "bbl"),
    ("BIN", "bin"),
    ("Landmark Status", "landmark_status"),
    ("Families", "families"),
    ("Violations", "violations"),
    ("CO", "co"),
    ("CO Date", "co_date"),
    ("Historical Image Cards", "historical_image_cards"),
    ("Historical Image Cards Date", "historical_image_cards_date"),
    ("Land Use", "land_use"),
    ("Lot Area (sq ft)", "lot_area"),
    ("Lot Frontage (ft)", "lot_frontage"),
    ("Lot Depth (ft)", "lot_depth"),
    ("Year Built", "year_built"),
    ("Year Altered", "year_altered"),
    ("Building Class", "building_class"),
    ("Zoning Districts", "zoning_districts"),
    ("HPD URL", "hpd_url"),
    ("DOB URL", "dob_url"),
    ("Notes", "notes"),
    ("CO Download Status", "co_download_status"),
    ("CO Retry URL", "co_retry_url"),
]


REPORT_ROWS = [
    ("Other address", "matched_address"),
    ("Landmark Status", "landmark_status"),
    ("Families", "families"),
    ("Violations", "violations"),
    ("CO", "co"),
    ("CO date", "co_date"),
    ("Historical Image cards", "historical_image_cards"),
    ("Historical Image cards date", "historical_image_cards_date"),
    ("Land Use", "land_use"),
    ("Lot Area", "lot_area"),
    ("Lot Frontage", "lot_frontage"),
    ("Lot Depth", "lot_depth"),
    ("Year Built", "year_built"),
    ("Year Altered", "year_altered"),
    ("Building Class", "building_class"),
    ("BBL", "bbl"),
    ("BIN", "bin"),
    ("Query Status", "status"),
    ("Notes / Error", "notes_or_error"),
    ("CO download status", "co_download_status"),
    ("CO retry URL", "co_retry_url"),
]


def read_addresses(content: bytes, requested_column: str = "Jobsite") -> tuple[list[str], str]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"无法读取 Excel：{exc}") from exc

    target = (requested_column or "Jobsite").strip().casefold()
    for sheet in workbook.worksheets:
        # Some Google Sheets exports omit the worksheet dimension metadata, so
        # openpyxl reports max_row/max_column as None.  Limit the iterator
        # directly instead of relying on those optional metadata values.
        preview_rows = itertools.islice(sheet.iter_rows(min_row=1, values_only=True), 50)
        for row_number, row in enumerate(preview_rows, start=1):
            headers = [str(value).strip().casefold() if value is not None else "" for value in row]
            if target in headers:
                column_number = headers.index(target) + 1
                addresses = []
                seen = set()
                for values in sheet.iter_rows(min_row=row_number + 1, min_col=column_number, max_col=column_number, values_only=True):
                    value = str(values[0]).strip() if values[0] is not None else ""
                    if value and value.casefold() not in seen:
                        seen.add(value.casefold())
                        addresses.append(value)
                return addresses, f"{sheet.title}!{get_column_letter(column_number)}{row_number}"
    raise ValueError(f"前 50 行中找不到列名“{requested_column or 'Jobsite'}”。")


def _safe_folder(address: str, index: int) -> str:
    value = re.sub(r"[<>:\\|?*\x00-\x1f]", "_", address).strip(" .")
    return f"{index:03d}_{value[:80] or 'address'}"


def _style_header(cells: Iterable) -> None:
    for cell in cells:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(vertical="center")


def build_workbook(records: list[PropertyRecord], source: str) -> bytes:
    workbook = Workbook()
    details = workbook.active
    details.title = "Results"
    details.append([label for label, _ in DETAIL_COLUMNS])
    _style_header(details[1])
    for record in records:
        data = record.as_dict()
        details.append([data.get(key, "") for _, key in DETAIL_COLUMNS])
    details.freeze_panes = "A2"
    details.auto_filter.ref = details.dimensions
    details.row_dimensions[1].height = 28
    for index, (label, _) in enumerate(DETAIL_COLUMNS, start=1):
        width = 14
        if "Address" in label or "URL" in label or label in {"Notes", "Error"}:
            width = 34
        details.column_dimensions[get_column_letter(index)].width = width
        for cell in details[get_column_letter(index)]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    report = workbook.create_sheet("Report View")
    report.cell(1, 1, "Field")
    for column, record in enumerate(records, start=2):
        report.cell(1, column, record.input_address)
    _style_header(report[1])
    for row, (label, key) in enumerate(REPORT_ROWS, start=2):
        report.cell(row, 1, label)
        report.cell(row, 1).font = Font(bold=True)
        report.cell(row, 1).fill = PatternFill("solid", fgColor="D9EAF7")
        for column, record in enumerate(records, start=2):
            data = record.as_dict()
            value = (data.get("notes") or data.get("error")) if key == "notes_or_error" else data.get(key, "")
            report.cell(row, column, value)
    report.freeze_panes = "B2"
    report.column_dimensions["A"].width = 30
    for column in range(2, len(records) + 2):
        report.column_dimensions[get_column_letter(column)].width = 35
    for row in report.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    errors = workbook.create_sheet("Errors")
    errors.append(["Input Address", "Status", "Error", "Notes"])
    _style_header(errors[1])
    for record in records:
        if record.status != "OK" or record.error or record.notes:
            errors.append([record.input_address, record.status, record.error, "; ".join(record.notes)])
    errors.freeze_panes = "A2"
    for col, width in zip("ABCD", [40, 14, 55, 70]):
        errors.column_dimensions[col].width = width

    about = workbook.create_sheet("About")
    about.append(["Item", "Value"])
    about.append(["Input source", source])
    about.append(["Generated fields", "NYC GeoSearch, MapPLUTO, HPD Open Data, DOB Open Data/BIS"])
    about.append(["Important", "Results are research aids. Verify legal occupancy and official documents with NYC agencies."])
    about.column_dimensions["A"].width = 22
    about.column_dimensions["B"].width = 95
    _style_header(about[1])
    for row in about.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_result_zip(
    records: list[PropertyRecord],
    documents: dict[str, list[tuple[str, bytes]]],
    source: str,
) -> bytes:
    output = io.BytesIO()
    failed_co_records: list[PropertyRecord] = []
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("property_results.xlsx", build_workbook(records, source))
        for index, record in enumerate(records, start=1):
            folder = _safe_folder(record.input_address, index)
            record_documents = documents.get(record.input_address, [])
            data = record.as_dict()
            information = [f"{label}: {data.get(key, '')}" for label, key in DETAIL_COLUMNS]
            if not record_documents:
                information.extend([
                    "",
                    "Downloaded documents: None",
                    "I-Card and CO PDFs are included only when the NYC agency site provides a directly downloadable PDF.",
                ])
            archive.writestr(
                f"information/{folder}/property_information.txt",
                "\n".join(information).encode("utf-8-sig"),
            )
            for filename, content in record_documents:
                archive.writestr(f"documents/{folder}/{Path(filename).name}", content)
            if record.bin and record.co_download_status != "Downloaded":
                failed_co_records.append(record)

        if failed_co_records:
            report = io.StringIO(newline="")
            writer = csv.writer(report)
            writer.writerow(["Input Address", "BIN", "Status", "Manual BIS URL", "Local Retry URL"])
            for record in failed_co_records:
                retry_url = "http://127.0.0.1:5000/retry-co?" + urlencode({
                    "bin": record.bin,
                    "address": record.input_address,
                })
                writer.writerow([
                    record.input_address,
                    record.bin,
                    record.co_download_status,
                    record.co_retry_url,
                    retry_url,
                ])
            archive.writestr("failed_co_downloads.csv", report.getvalue().encode("utf-8-sig"))
    return output.getvalue()
